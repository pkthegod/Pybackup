import os
import re
import multiprocessing
from typing import Set, List, Iterator, Tuple
import logging
from concurrent.futures import ProcessPoolExecutor
from functools import lru_cache

import fitz  # PyMuPDF
from openpyxl import load_workbook
from tqdm import tqdm

# Configurações de Validação de Domínio
MIN_DOMAIN_LENGTH = 4
MAX_DOMAIN_LENGTH = 253

# Regex otimizada para captura de domínios
DOMAIN_REGEX = re.compile(
    r'\b((?!-)[A-Za-z0-9-]{1,63}(?<!-)\.)+(?!-)(?:[A-Za-z]{2,63})(?<!-)\b',
    re.IGNORECASE
)

# Configurar logging para exibir no terminal
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Usar cache para evitar revalidar domínios repetidos
@lru_cache(maxsize=10000)
def validate_domain(domain: str) -> bool:
    """Valida um domínio seguindo regras ICANN com cache para melhor desempenho."""
    if len(domain) < MIN_DOMAIN_LENGTH or len(domain) > MAX_DOMAIN_LENGTH:
        return False
    
    parts = domain.split('.')
    if len(parts) < 2:
        return False
    
    tld = parts[-1]
    if not tld.isalpha() or len(tld) < 2:
        return False
    
    for part in parts:
        if not (1 <= len(part) <= 63):
            return False
        if part.startswith('-') or part.endswith('-'):
            return False
        if not all(c.isalnum() or c == '-' for c in part):
            return False
    
    return True

def clean_domain(candidate: str) -> str | None:
    """Limpa e normaliza um domínio antes da validação."""
    try:
        if '.' not in candidate:
            return None

        candidate = candidate.split('//')[-1].split('/')[0].split('?')[0]
        candidate = candidate.split(':')[0].lower().strip('.:,;!*')
        
        if '.' not in candidate:
            return None
        
        parts = candidate.split('.')
        if not parts[-1].isalpha():
            return None
        
        return candidate if validate_domain(candidate) else None
    except Exception:
        return None

def find_domains(text: str) -> Set[str]:
    """Procura e valida domínios no texto de forma otimizada."""
    if not text:
        return set()
    
    matches = DOMAIN_REGEX.finditer(text)
    candidates = (match.group() for match in matches)
    
    return {domain for domain in (clean_domain(candidate) for candidate in candidates) if domain}

def extract_text_from_pdf(pdf_path: str) -> Iterator[str]:
    """Extrai texto de PDFs página por página para economizar memória."""
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                yield page.get_text("text")
    except Exception as e:
        logging.error(f"Erro no PDF {pdf_path}: {str(e)}")

def extract_text_from_excel(excel_path: str) -> Iterator[str]:
    """Extrai texto de arquivos Excel de forma otimizada."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        for sheet in wb:
            buffer = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text:
                    buffer.append(row_text)
                
                if len(buffer) >= 100:
                    yield "\n".join(buffer)
                    buffer = []
            
            if buffer:
                yield "\n".join(buffer)
    except Exception as e:
        logging.error(f"Erro no Excel {excel_path}: {str(e)}")

def process_single_file(file_path: str) -> Set[str]:
    """Processa um único arquivo e extrai domínios de forma otimizada."""
    try:
        extension = os.path.splitext(file_path)[1].lower()
        
        if extension == '.pdf':
            domains = set()
            for page_text in extract_text_from_pdf(file_path):
                domains.update(find_domains(page_text))
        elif extension in ('.xls', '.xlsx', '.ods'):
            domains = set()
            for text_batch in extract_text_from_excel(file_path):
                domains.update(find_domains(text_batch))
        else:
            return set()
        
        if domains:
            file_name = os.path.basename(file_path)
            logging.info(f"✅ Domínios encontrados em {file_name}: {len(domains)}")
        
        return domains
    
    except Exception as e:
        file_name = os.path.basename(file_path)
        logging.error(f"Erro ao processar {file_name}: {str(e)}")
        return set()

def get_file_size_category(file_path: str) -> int:
    """Categoriza arquivos por tamanho para otimizar processamento."""
    try:
        size = os.path.getsize(file_path)
        if size > 50 * 1024 * 1024:  # Maior que 50MB
            return 3
        elif size > 10 * 1024 * 1024:  # Entre 10MB e 50MB
            return 2
        elif size > 1 * 1024 * 1024:  # Entre 1MB e 10MB
            return 1
        else:  # Menor que 1MB
            return 0
    except:
        return 0

def process_files(folder_path: str, max_workers: int = None) -> Set[str]:
    """Processa arquivos em paralelo com otimização baseada em tamanho."""
    all_files = [
        os.path.join(root, f)
        for root, _, files in os.walk(folder_path)
        for f in files if os.path.splitext(f)[1].lower() in {'.pdf', '.xls', '.xlsx', '.ods'}
    ]
    
    if not all_files:
        logging.warning("⚠️ Nenhum arquivo compatível encontrado na pasta.")
        return set()
    
    max_workers = max_workers or max(1, min(os.cpu_count() or 1, 8))
    
    files_by_size = {}
    for file_path in all_files:
        category = get_file_size_category(file_path)
        if category not in files_by_size:
            files_by_size[category] = []
        files_by_size[category].append(file_path)
    
    all_domains = set()
    
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        if 3 in files_by_size:
            large_files = files_by_size[3]
            for result in tqdm(
                executor.map(process_single_file, large_files),
                total=len(large_files),
                desc="Processando arquivos grandes"
            ):
                all_domains.update(result)
        
        medium_files = []
        for category in [1, 2]:
            if category in files_by_size:
                medium_files.extend(files_by_size[category])
        
        if medium_files:
            batch_size = max(1, len(medium_files) // (max_workers * 2))
            futures = []
            for i in range(0, len(medium_files), batch_size):
                batch = medium_files[i:i+batch_size]
                for file in batch:
                    futures.append(executor.submit(process_single_file, file))
            
            for future in tqdm(
                futures,
                total=len(medium_files),
                desc="Processando arquivos médios"
            ):
                all_domains.update(future.result())
        
        if 0 in files_by_size:
            small_files = files_by_size[0]
            batch_size = max(10, len(small_files) // max_workers)
            
            for result in tqdm(
                executor.map(process_single_file, small_files, chunksize=batch_size),
                total=len(small_files),
                desc="Processando arquivos pequenos"
            ):
                all_domains.update(result)
    
    return all_domains

def read_existing_domains(file_path: str) -> Set[str]:
    """Lê domínios de um arquivo de texto de forma otimizada."""
    if not os.path.exists(file_path):
        logging.warning(f"⚠️ Arquivo {file_path} não encontrado.")
        return set()
    
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return {line.strip() for line in f if line.strip()}
    except Exception as e:
        logging.error(f"Erro ao ler arquivo de domínios: {str(e)}")
        return set()

def main(input_folder: str = r"C:\pdfs\liberados", output_file: str = "dominios_liberados.txt"):
    """Função principal para extração, combinação e salvamento de domínios."""
    logging.info("Iniciando processamento otimizado...")
    
    try:
        if not os.path.exists(input_folder):
            logging.error(f"⚠️ Pasta '{input_folder}' não encontrada!")
            return
        
        extracted_domains = process_files(input_folder)
        
        if not extracted_domains:
            logging.info("⚠️ Nenhum domínio extraído dos arquivos.")
        
        existing_domains_path = os.path.join(input_folder, "dominios_liberados.txt")
        existing_domains = read_existing_domains(existing_domains_path)
        
        combined_domains = sorted(extracted_domains.union(existing_domains))
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(combined_domains))
        
        logging.info(f"\n✅ Concluído! Total de domínios únicos salvos: {len(combined_domains)}")
        
        new_domains = len(extracted_domains - existing_domains)
        if new_domains > 0:
            logging.info(f"✅ Novos domínios encontrados: {new_domains}")
    
    except Exception as e:
        logging.error(f"Erro ao processar ou salvar resultados: {str(e)}")

if __name__ == '__main__':
    main()