import re
import os
import fitz  # PyMuPDF
import pandas as pd
from concurrent.futures import ProcessPoolExecutor
from openpyxl import load_workbook
from tqdm import tqdm

# Configurações Gerais
MIN_DOMAIN_LENGTH = 4
MAX_DOMAIN_LENGTH = 253

# Regex otimizada para captura de domínios
DOMAIN_REGEX = re.compile(
    r'\b((?!-)[A-Za-z0-9-]{1,63}(?<!-)\.)+(?!-)[A-Za-z0-9-]{1,63}(?<!-)\b',
    re.IGNORECASE
)

def validate_domain(domain):
    """
    Validação rigorosa de domínios de acordo com padrões ICANN.
    Além de checar comprimentos e caracteres válidos, garante que o TLD seja
    composto apenas por letras.
    """
    if len(domain) < MIN_DOMAIN_LENGTH or len(domain) > MAX_DOMAIN_LENGTH:
        return False

    parts = domain.split('.')
    # Deve ter pelo menos duas partes
    if len(parts) < 2:
        return False

    # Validação do TLD: somente letras são permitidas
    tld = parts[-1]
    if not tld.isalpha() or len(tld) < 2:
        return False

    # Validação de cada parte do domínio
    for part in parts:
        if len(part) < 1 or len(part) > 63:
            return False
        # Não pode começar ou terminar com hífen
        if part.startswith('-') or part.endswith('-'):
            return False
        # Somente letras, números e hífens
        if not re.match(r'^[a-z0-9-]+$', part, re.IGNORECASE):
            return False

    return True

def clean_domain(candidate):
    """
    Limpa e normaliza candidatos a domínio e valida seu formato.
    Remove protocolos, caminhos, portas e parâmetros, deixando apenas o domínio.
    Retorna o domínio limpo se for válido, ou None caso contrário.
    """
    # Remove protocolos (http, https) e extrai apenas a parte do domínio
    candidate = candidate.split('//')[-1].split('/')[0].split('?')[0]
    # Remove portas, vírgulas e demais pontuações indesejadas
    candidate = candidate.split(':')[0].lower().strip('.:,;!*')
    
    # Verifica se o candidato possui pelo menos um ponto
    if '.' not in candidate:
        return None

    # Garante que o TLD (parte após o último ponto) contenha somente letras
    parts = candidate.split('.')
    if not parts[-1].isalpha():
        return None

    # Validação final segundo regras ICANN
    return candidate if validate_domain(candidate) else None

def find_domains(text):
    """
    Procura por domínios no texto utilizando regex e filtragem.
    Apenas adiciona domínios que passam pelo processo de limpeza e validação.
    """
    found = set()
    if not text:
        return found
    
    for match in DOMAIN_REGEX.finditer(text):
        candidate = match.group()
        domain = clean_domain(candidate)
        if domain:
            found.add(domain)
    return found

def extract_text_from_pdf(pdf_path):
    """
    Extrai texto de PDFs utilizando blocos (apenas o conteúdo textual).
    """
    text_blocks = []
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                blocks = page.get_text("blocks", sort=True)
                # block[4] contém o texto, e block[6] indica o tipo (0 para texto)
                text_blocks.extend(block[4] for block in blocks if block[6] == 0)
        return "\n".join(text_blocks)
    except Exception as e:
        print(f"Erro no PDF {pdf_path}: {str(e)}")
        return ""

def extract_text_from_excel(excel_path):
    """
    Extrai texto de planilhas utilizando openpyxl para maior eficiência.
    """
    texts = []
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for sheet in wb:
            for row in sheet.iter_rows(values_only=True):
                texts.extend(str(cell) for cell in row if cell)
        return "\n".join(texts)
    except Exception as e:
        print(f"Erro no Excel {excel_path}: {str(e)}")
        return ""

def process_single_file(file_path):
    """
    Processa um único arquivo (PDF ou planilha) e extrai os domínios encontrados.
    """
    try:
        if file_path.lower().endswith('.pdf'):
            text = extract_text_from_pdf(file_path)
        elif file_path.lower().endswith(('.xls', '.xlsx', '.ods')):
            text = extract_text_from_excel(file_path)
        else:
            return set()
        
        return find_domains(text)
    except Exception as e:
        print(f"Erro em {os.path.basename(file_path)}: {str(e)}")
        return set()

def process_files(folder_path, workers=None):
    """
    Processa arquivos em paralelo, com uma barra de progresso para acompanhar.
    Procura por arquivos PDF e planilhas, extraindo domínios únicos de cada um.
    """
    all_files = []
    for root, _, files in os.walk(folder_path):
        all_files.extend(
            os.path.join(root, f)
            for f in files if f.lower().endswith(('.pdf', '.xls', '.xlsx', '.ods'))
        )
    
    all_domains = set()
    with ProcessPoolExecutor(max_workers=workers or os.cpu_count()) as executor:
        results = list(tqdm(
            executor.map(process_single_file, all_files),
            total=len(all_files),
            desc="Processando arquivos"
        ))
    
    for domains in results:
        all_domains.update(domains)
    
    return sorted(all_domains)

# Configurações de entrada e saída
input_folder = r"C:\pdfs"  # Ajuste conforme sua pasta de arquivos
output_file = "dominios_extraidos.txt"

if __name__ == '__main__':
    print("Iniciando processamento paralelo...")
    result = process_files(input_folder)
    
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(result))
        print(f"\n✅ Concluído! Total de domínios únicos encontrados: {len(result)}")
    except Exception as e:
        print(f"Erro ao salvar resultados: {str(e)}")
