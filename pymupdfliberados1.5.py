import os
import re
import multiprocessing
from typing import Set, List, Iterator
import logging
from concurrent.futures import ProcessPoolExecutor
from functools import lru_cache

import fitz  # PyMuPDF
from openpyxl import load_workbook
from tqdm import tqdm

# Configurações
MIN_DOMAIN_LENGTH = 4
MAX_DOMAIN_LENGTH = 253
DOMAIN_REGEX = re.compile(
    r'\b((?!-)[A-Za-z0-9-]{1,63}(?<!-)\.)+(?!-)(?:[A-Za-z]{2,63})(?<!-)\b',
    re.IGNORECASE
)
FIXED_DOMAINS = {
    "exemplo1.com",
    "exemplo2.org",
    "exemplo3.net",
    # Adicione mais domínios aqui
}

# Configuração de logging
logging.basicConfig(level=logging.INFO, format='%(message)s')


@lru_cache(maxsize=10000)
def validate_domain(domain: str) -> bool:
    """Valida um domínio seguindo as regras ICANN."""
    if not (MIN_DOMAIN_LENGTH <= len(domain) <= MAX_DOMAIN_LENGTH):
        return False

    parts = domain.split('.')
    if len(parts) < 2:
        return False

    tld = parts[-1]
    if not tld.isalpha() or len(tld) < 2:
        return False

    for part in parts:
        if not (1 <= len(part) <= 63):
            return False
        if part.startswith('-') or part.endswith('-'):
            return False
        if not all(c.isalnum() or c == '-' for c in part):
            return False

    return True


def clean_domain(candidate: str) -> str | None:
    """Limpa e normaliza um domínio antes da validação."""
    if '.' not in candidate:
        return None

    candidate = candidate.split('//')[-1].split('/')[0].split('?')[0]
    candidate = candidate.split(':')[0].lower().strip('.:,;!*')

    if '.' not in candidate:
        return None

    parts = candidate.split('.')
    if not parts[-1].isalpha():
        return None

    return candidate if validate_domain(candidate) else None


def find_domains(text: str) -> Set[str]:
    """Procura e valida domínios no texto."""
    if not text:
        return set()

    matches = DOMAIN_REGEX.finditer(text)
    candidates = (match.group() for match in matches)
    return {domain for candidate in candidates if (domain := clean_domain(candidate))}


def extract_text_from_pdf(pdf_path: str) -> Iterator[str]:
    """Extrai texto de um arquivo PDF com verificação de encoding."""
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                text = page.get_text("text", flags=fitz.TEXT_PRESERVE_WHITESPACE)
                if text:
                    # Normaliza espaços e quebras de linha
                    text = ' '.join(text.split()).replace('\\n', ' ')
                    yield text
                else:
                    logging.warning(f"PDF sem texto legível: {pdf_path}")
    except Exception as e:
        logging.error(f"Erro crítico no PDF {pdf_path}: {str(e)}")


def extract_text_from_excel(excel_path: str) -> Iterator[str]:
    """Extrai texto de um arquivo Excel."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for sheet in wb:
            buffer = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join(str(cell) for cell in row if cell is not None)
                if row_text:
                    buffer.append(row_text)

                if len(buffer) >= 100:
                    yield "\n".join(buffer)
                    buffer = []

            if buffer:
                yield "\n".join(buffer)
    except Exception as e:
        logging.error(f"Erro ao processar Excel {excel_path}: {e}")


def process_single_file(file_path: str) -> Set[str]:
    """Processa um único arquivo e extrai domínios."""
    try:
        extension = os.path.splitext(file_path)[1].lower()

        if extension == '.pdf':
            domains = set()
            for page_text in extract_text_from_pdf(file_path):
                domains.update(find_domains(page_text))
        elif extension in ('.xls', '.xlsx', '.ods'):
            domains = set()
            for text_batch in extract_text_from_excel(file_path):
                domains.update(find_domains(text_batch))
        else:
            return set()

        if domains:
            file_name = os.path.basename(file_path)
            logging.info(f"✅ Domínios encontrados em {file_name}: {len(domains)}")

        return domains

    except Exception as e:
        file_name = os.path.basename(file_path)
        logging.error(f"Erro ao processar {file_name}: {e}")
        return set()


def get_file_size_category(file_path: str) -> int:
    """Categoriza arquivos por tamanho para otimização."""
    try:
        size = os.path.getsize(file_path)
        if size > 50 * 1024 * 1024:  # > 50MB
            return 3
        elif size > 10 * 1024 * 1024:  # 10MB - 50MB
            return 2
        elif size > 1 * 1024 * 1024:  # 1MB - 10MB
            return 1
        else:  # < 1MB
            return 0
    except Exception:
        return 0


def process_files(folder_path: str, max_workers: int = None) -> Set[str]:
    """Processa arquivos em paralelo, otimizado por tamanho."""
    all_files = [
        os.path.join(root, f)
        for root, _, files in os.walk(folder_path)
        for f in files if os.path.splitext(f)[1].lower() in {'.pdf', '.xls', '.xlsx', '.ods'}
    ]

    if not all_files:
        logging.warning("⚠️ Nenhum arquivo compatível encontrado na pasta.")
        return set()

    max_workers = max_workers or max(1, min(os.cpu_count() or 1, 8))
    files_by_size = {0: [], 1: [], 2: [], 3: []}

    for file_path in all_files:
        category = get_file_size_category(file_path)
        files_by_size[category].append(file_path)

    all_domains = set()

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # Processar arquivos grandes
        if files_by_size[3]:
            for result in tqdm(
                executor.map(process_single_file, files_by_size[3]),
                total=len(files_by_size[3]),
                desc="Processando arquivos grandes"
            ):
                all_domains.update(result)

        # Processar arquivos médios
        medium_files = files_by_size[1] + files_by_size[2]
        if medium_files:
            for result in tqdm(
                executor.map(process_single_file, medium_files),
                total=len(medium_files),
                desc="Processando arquivos médios"
            ):
                all_domains.update(result)

        # Processar arquivos pequenos
        if files_by_size[0]:
            for result in tqdm(
                executor.map(process_single_file, files_by_size[0], chunksize=10),
                total=len(files_by_size[0]),
                desc="Processando arquivos pequenos"
            ):
                all_domains.update(result)

    return all_domains


def read_existing_domains(file_path: str) -> Set[str]:
    """Lê domínios de um arquivo de texto."""
    if not os.path.exists(file_path):
        logging.warning(f"⚠️ Arquivo {file_path} não encontrado.")
        return set()

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return {line.strip() for line in f if line.strip()}
    except Exception as e:
        logging.error(f"Erro ao ler arquivo de domínios: {e}")
        return set()


def save_domains(output_file: str, domains: Set[str]):
    """Salva domínios em um arquivo de texto."""
    try:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(sorted(domains)))
        logging.info(f"✅ Domínios salvos em {output_file}")
    except Exception as e:
        logging.error(f"Erro ao salvar domínios: {e}")


def main(input_folder: str = r"C:\pdfs\liberados", output_file: str = "dominios_liberados.txt"):
    """Função principal para extração, combinação e salvamento de domínios."""
    logging.info("Iniciando processamento...")

    if not os.path.exists(input_folder):
        logging.error(f"⚠️ Pasta '{input_folder}' não encontrada!")
        return

    # Verifica todos os arquivos na pasta
    all_files = [
        os.path.join(root, f)
        for root, _, files in os.walk(input_folder)
        for f in files if os.path.splitext(f)[1].lower() in {'.pdf', '.xls', '.xlsx', '.ods'}
    ]

    if not all_files:
        logging.warning("⚠️ Nenhum arquivo compatível encontrado na pasta.")
        return

    logging.info(f"🔍 Total de arquivos encontrados: {len(all_files)}")
    for file_path in all_files:
        logging.info(f"📄 Arquivo: {file_path}")

    extracted_domains = process_files(input_folder)
    existing_domains = read_existing_domains(os.path.join(input_folder, "dominios_liberados.txt"))

    # Combina domínios extraídos, existentes e fixos
    combined_domains = extracted_domains.union(existing_domains).union(FIXED_DOMAINS)
    save_domains(output_file, combined_domains)

    logging.info(f"✅ Concluído! Total de domínios únicos: {len(combined_domains)}")
    if new_domains := len(extracted_domains - existing_domains):
        logging.info(f"✅ Novos domínios encontrados: {new_domains}")


if __name__ == '__main__':
    main()