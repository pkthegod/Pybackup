import os
import re
import multiprocessing
from typing import Set, List

import fitz  # PyMuPDF
from openpyxl import load_workbook
from tqdm import tqdm

# Configurações de Validação de Domínio
MIN_DOMAIN_LENGTH = 4
MAX_DOMAIN_LENGTH = 253

# Regex otimizada para captura de domínios
DOMAIN_REGEX = re.compile(
    r'\b((?!-)[A-Za-z0-9-]{1,63}(?<!-)\.)+(?!-)[A-Za-z0-9-]{1,63}(?<!-)\b',
    re.IGNORECASE
)

def validate_domain(domain: str) -> bool:
    """
    Valida um domínio seguindo regras ICANN com verificações mais eficientes.
    
    Args:
        domain (str): Domínio a ser validado
    
    Returns:
        bool: Indica se o domínio é válido
    """
    # Verificações iniciais rápidas
    if len(domain) < MIN_DOMAIN_LENGTH or len(domain) > MAX_DOMAIN_LENGTH:
        return False
    
    parts = domain.split('.')
    if len(parts) < 2:
        return False
    
    # Validação do TLD
    tld = parts[-1]
    if not tld.isalpha() or len(tld) < 2:
        return False
    
    # Validação de cada parte do domínio
    for part in parts:
        # Comprimento da parte do domínio
        if not (1 <= len(part) <= 63):
            return False
        
        # Verifica caracteres inválidos e hífens
        if part.startswith('-') or part.endswith('-'):
            return False
        
        if not all(c.isalnum() or c == '-' for c in part):
            return False
    
    return True

def clean_domain(candidate: str) -> str | None:
    """
    Limpa e normaliza um domínio antes da validação.
    
    Args:
        candidate (str): Domínio candidato
    
    Returns:
        str | None: Domínio limpo ou None se inválido
    """
    # Remoção de prefixos, portas, parâmetros
    candidate = candidate.split('//')[-1].split('/')[0].split('?')[0]
    candidate = candidate.split(':')[0].lower().strip('.:,;!*')
    
    # Verificação de TLD
    if '.' not in candidate:
        return None
    
    parts = candidate.split('.')
    if not parts[-1].isalpha():
        return None
    
    return candidate if validate_domain(candidate) else None

def find_domains(text: str) -> Set[str]:
    """
    Procura e valida domínios no texto.
    
    Args:
        text (str): Texto para busca de domínios
    
    Returns:
        Set[str]: Conjunto de domínios válidos encontrados
    """
    if not text:
        return set()
    
    found = set()
    for match in DOMAIN_REGEX.finditer(text):
        domain = clean_domain(match.group())
        if domain:
            found.add(domain)
    
    return found

def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Extração de texto de PDFs de forma mais leve.
    
    Args:
        pdf_path (str): Caminho para o arquivo PDF
    
    Returns:
        str: Texto extraído do PDF
    """
    try:
        with fitz.open(pdf_path) as doc:
            return " ".join(page.get_text("text") for page in doc)
    except Exception as e:
        print(f"Erro no PDF {pdf_path}: {str(e)}")
        return ""

def extract_text_from_excel(excel_path: str) -> str:
    """
    Extração de texto de arquivos Excel de forma otimizada.
    
    Args:
        excel_path (str): Caminho para o arquivo Excel
    
    Returns:
        str: Texto extraído do arquivo Excel
    """
    try:
        texts = []
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for sheet in wb:
            texts.extend(
                str(cell) for row in sheet.iter_rows(values_only=True)
                for cell in row if cell is not None
            )
        return "\n".join(texts)
    except Exception as e:
        print(f"Erro no Excel {excel_path}: {str(e)}")
        return ""

def process_single_file(file_path: str) -> Set[str]:
    """
    Processa um único arquivo e extrai domínios.
    
    Args:
        file_path (str): Caminho do arquivo
    
    Returns:
        Set[str]: Domínios encontrados no arquivo
    """
    try:
        # Seleção do método de extração baseado na extensão
        extractors = {
            '.pdf': extract_text_from_pdf,
            '.xls': extract_text_from_excel,
            '.xlsx': extract_text_from_excel,
            '.ods': extract_text_from_excel
        }
        
        # Obtém extrator baseado na extensão
        extractor = extractors.get(os.path.splitext(file_path)[1].lower())
        
        if not extractor:
            return set()
        
        # Extrai texto e encontra domínios
        text = extractor(file_path)
        domains = find_domains(text)
        
        if domains:
            print(f"✅ Domínios encontrados em {os.path.basename(file_path)}: {len(domains)}")
        
        return domains
    
    except Exception as e:
        print(f"Erro ao processar {os.path.basename(file_path)}: {str(e)}")
        return set()

def process_files(
    folder_path: str, 
    batch_size: int = None, 
    max_workers: int = None
) -> List[str]:
    """
    Processa arquivos em paralelo com gerenciamento eficiente de recursos.
    
    Args:
        folder_path (str): Diretório com arquivos
        batch_size (int, optional): Tamanho do lote de processamento
        max_workers (int, optional): Número máximo de workers
    
    Returns:
        List[str]: Lista ordenada de domínios encontrados
    """
    # Coleta de arquivos elegíveis
    all_files = [
        os.path.join(root, f)
        for root, _, files in os.walk(folder_path)
        for f in files if os.path.splitext(f)[1].lower() in {'.pdf', '.xls', '.xlsx', '.ods'}
    ]
    
    # Ajuste automático de workers
    max_workers = max_workers or (os.cpu_count() or 1)
    
    # Ajuste de batch size para reduzir overhead
    if batch_size is None:
        batch_size = max(1, len(all_files) // (max_workers * 2))
    
    all_domains = set()
    
    # Processamento em pool com imap_unordered
    with multiprocessing.Pool(processes=max_workers) as pool:
        for batch_result in tqdm(
            pool.imap_unordered(process_single_file, all_files, chunksize=batch_size),
            total=len(all_files),
            desc="Processando arquivos"
        ):
            all_domains.update(batch_result)
    
    return sorted(all_domains)

def main(input_folder: str = r"C:\pdfs", output_file: str = "dominios_extraidos.txt"):
    """
    Função principal para extração de domínios.
    
    Args:
        input_folder (str): Diretório de entrada
        output_file (str): Arquivo de saída
    """
    print("Iniciando processamento paralelo...")
    
    try:
        result = process_files(input_folder)
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(result))
        
        print(f"\n✅ Concluído! Total de domínios únicos encontrados: {len(result)}")
    
    except Exception as e:
        print(f"Erro ao salvar resultados: {str(e)}")

if __name__ == '__main__':
    main()