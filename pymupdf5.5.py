import os
import re
import hashlib
import logging
from pathlib import Path
from typing import Set, List, Optional
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass
import fitz  # PyMuPDF
from openpyxl import load_workbook
from tqdm import tqdm


@dataclass
class ProcessingResult:
    """Resultado do processamento de um arquivo"""
    file_path: str
    domains: Set[str]
    success: bool
    error_message: Optional[str] = None


class DomainExtractor:
    """Classe principal para extração de domínios de arquivos"""
    
    # Configurações
    MIN_DOMAIN_LENGTH = 4
    MAX_DOMAIN_LENGTH = 253
    SUPPORTED_EXTENSIONS = {'.pdf', '.xls', '.xlsx', '.ods'}
    
    # Regex otimizada para captura de domínios
    DOMAIN_PATTERN = re.compile(
        r'\b(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}\b',
        re.IGNORECASE
    )
    
    def __init__(self, max_workers: Optional[int] = None):
        self.max_workers = max_workers or min(os.cpu_count() or 1, 8)
        self.setup_logging()
        self._domain_cache = {}
    
    def setup_logging(self):
        """Configura o sistema de logging"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(),
                logging.FileHandler('domain_extraction.log', encoding='utf-8')
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def validate_domain(self, domain: str) -> bool:
        """Valida se um domínio é válido seguindo as regras básicas"""
        if domain in self._domain_cache:
            return self._domain_cache[domain]
        
        is_valid = self._validate_domain_logic(domain)
        self._domain_cache[domain] = is_valid
        return is_valid
    
    def _validate_domain_logic(self, domain: str) -> bool:
        """Lógica de validação de domínio"""
        # Verificações básicas de comprimento
        if not (self.MIN_DOMAIN_LENGTH <= len(domain) <= self.MAX_DOMAIN_LENGTH):
            return False
        
        # Dividir em partes
        parts = domain.split('.')
        if len(parts) < 2:
            return False
        
        # Validar TLD (última parte)
        tld = parts[-1]
        if not tld.isalpha() or len(tld) < 2:
            return False
        
        # Validar cada parte do domínio
        for part in parts:
            if not (1 <= len(part) <= 63):
                return False
            if part.startswith('-') or part.endswith('-'):
                return False
            if not all(c.isalnum() or c == '-' for c in part):
                return False
        
        return True
    
    def clean_and_extract_domains(self, text: str) -> Set[str]:
        """Extrai e limpa domínios do texto"""
        if not text:
            return set()
        
        domains = set()
        
        # Encontrar todos os candidatos a domínio
        matches = self.DOMAIN_PATTERN.findall(text)
        
        for match in matches:
            # Limpar o domínio
            cleaned = self._clean_domain(match)
            if cleaned and self.validate_domain(cleaned):
                domains.add(cleaned.lower())
        
        return domains
    
    def _clean_domain(self, domain: str) -> Optional[str]:
        """Limpa um domínio candidato"""
        try:
            # Remove protocolos e caminhos
            if '//' in domain:
                domain = domain.split('//')[-1]
            
            # Remove paths, query strings, etc.
            domain = domain.split('/')[0].split('?')[0].split('#')[0]
            
            # Remove portas
            if ':' in domain and not domain.count(':') > 1:  # Não é IPv6
                domain = domain.split(':')[0]
            
            # Remove caracteres inválidos do início/fim
            domain = domain.strip('.,;:!?*()[]{}"\' \t\n\r')
            
            return domain if domain else None
        except Exception:
            return None
    
    def extract_from_pdf(self, file_path: str) -> Set[str]:
        """Extrai domínios de um arquivo PDF"""
        domains = set()
        try:
            with fitz.open(file_path) as doc:
                for page_num in range(len(doc)):
                    page = doc[page_num]
                    text = page.get_text()
                    page_domains = self.clean_and_extract_domains(text)
                    domains.update(page_domains)
                    
                    # Log progresso para arquivos grandes
                    if len(doc) > 50 and page_num % 20 == 0:
                        self.logger.debug(f"Processando página {page_num + 1}/{len(doc)} de {os.path.basename(file_path)}")
        
        except Exception as e:
            raise Exception(f"Erro ao processar PDF: {str(e)}")
        
        return domains
    
    def extract_from_excel(self, file_path: str) -> Set[str]:
        """Extrai domínios de um arquivo Excel"""
        domains = set()
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Processar em lotes para economia de memória
                for row in sheet.iter_rows(values_only=True):
                    row_text = ' '.join(str(cell) for cell in row if cell is not None)
                    row_domains = self.clean_and_extract_domains(row_text)
                    domains.update(row_domains)
            
            wb.close()
        
        except Exception as e:
            raise Exception(f"Erro ao processar Excel: {str(e)}")
        
        return domains
    
    def process_file(self, file_path: str) -> ProcessingResult:
        """Processa um único arquivo e retorna o resultado"""
        try:
            file_path = str(file_path)  # Garantir que é string
            extension = Path(file_path).suffix.lower()
            
            if extension == '.pdf':
                domains = self.extract_from_pdf(file_path)
            elif extension in {'.xls', '.xlsx', '.ods'}:
                domains = self.extract_from_excel(file_path)
            else:
                return ProcessingResult(
                    file_path=file_path,
                    domains=set(),
                    success=False,
                    error_message=f"Extensão não suportada: {extension}"
                )
            
            return ProcessingResult(
                file_path=file_path,
                domains=domains,
                success=True
            )
        
        except Exception as e:
            return ProcessingResult(
                file_path=file_path,
                domains=set(),
                success=False,
                error_message=str(e)
            )
    
    def find_files(self, folder_path: str) -> List[str]:
        """Encontra todos os arquivos suportados na pasta"""
        files = []
        folder = Path(folder_path)
        
        if not folder.exists():
            raise FileNotFoundError(f"Pasta não encontrada: {folder_path}")
        
        for file_path in folder.rglob("*"):
            if file_path.is_file() and file_path.suffix.lower() in self.SUPPORTED_EXTENSIONS:
                files.append(str(file_path))
        
        return files
    
    def remove_duplicates(self, folder_path: str) -> int:
        """Remove arquivos duplicados baseado no hash MD5"""
        self.logger.info("Verificando arquivos duplicados...")
        
        file_hashes = {}
        files_to_remove = []
        
        # Calcular hashes
        for file_path in self.find_files(folder_path):
            try:
                file_hash = self._calculate_file_hash(file_path)
                if file_hash in file_hashes:
                    files_to_remove.append(file_path)
                    self.logger.info(f"Duplicado encontrado: {os.path.basename(file_path)}")
                else:
                    file_hashes[file_hash] = file_path
            except Exception as e:
                self.logger.error(f"Erro ao calcular hash de {file_path}: {e}")
        
        # Remover duplicados
        removed_count = 0
        for file_path in files_to_remove:
            try:
                os.remove(file_path)
                removed_count += 1
                self.logger.info(f"Arquivo removido: {os.path.basename(file_path)}")
            except Exception as e:
                self.logger.error(f"Erro ao remover {file_path}: {e}")
        
        return removed_count
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calcula o hash MD5 de um arquivo"""
        hasher = hashlib.md5()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b""):
                hasher.update(chunk)
        return hasher.hexdigest()
    
    def load_existing_domains(self, file_path: str) -> Set[str]:
        """Carrega domínios existentes de um arquivo"""
        domains = set()
        
        if not os.path.exists(file_path):
            self.logger.info(f"Arquivo de domínios existentes não encontrado: {file_path}")
            return domains
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    domain = line.strip()
                    if domain and self.validate_domain(domain):
                        domains.add(domain.lower())
            
            self.logger.info(f"Carregados {len(domains)} domínios existentes")
        
        except Exception as e:
            self.logger.error(f"Erro ao carregar domínios existentes: {e}")
        
        return domains
    
    def save_domains(self, domains: Set[str], output_file: str) -> bool:
        """Salva os domínios em um arquivo"""
        try:
            # Criar diretório se não existir
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Ordenar domínios para facilitar leitura
            sorted_domains = sorted(domains)
            
            # Salvar no arquivo
            with open(output_file, 'w', encoding='utf-8') as f:
                for domain in sorted_domains:
                    f.write(f"{domain}\n")
            
            self.logger.info(f"✅ Domínios salvos em: {output_file}")
            self.logger.info(f"✅ Total de domínios únicos: {len(sorted_domains)}")
            return True
        
        except Exception as e:
            self.logger.error(f"❌ Erro ao salvar domínios: {e}")
            return False
    
    def extract_domains(self, input_folder: str, output_file: str = "dominios_extraidos.txt", 
                       remove_duplicates: bool = True, existing_domains_file: Optional[str] = None) -> bool:
        """Função principal para extrair domínios"""
        try:
            self.logger.info("🚀 Iniciando extração de domínios...")
            
            # Remover duplicados se solicitado
            if remove_duplicates:
                removed = self.remove_duplicates(input_folder)
                if removed > 0:
                    self.logger.info(f"🗑️ Removidos {removed} arquivos duplicados")
            
            # Encontrar arquivos
            files = self.find_files(input_folder)
            if not files:
                self.logger.warning("⚠️ Nenhum arquivo compatível encontrado")
                return False
            
            self.logger.info(f"📁 Encontrados {len(files)} arquivos para processar")
            
            # Processar arquivos em paralelo
            all_domains = set()
            successful_files = 0
            failed_files = 0
            
            with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
                # Submeter todos os trabalhos
                future_to_file = {executor.submit(self.process_file, file): file for file in files}
                
                # Processar resultados com barra de progresso
                for future in tqdm(as_completed(future_to_file), total=len(files), desc="Processando arquivos"):
                    result = future.result()
                    
                    if result.success:
                        all_domains.update(result.domains)
                        successful_files += 1
                        if result.domains:
                            self.logger.info(f"✅ {os.path.basename(result.file_path)}: {len(result.domains)} domínios")
                    else:
                        failed_files += 1
                        self.logger.error(f"❌ {os.path.basename(result.file_path)}: {result.error_message}")
            
            # Carregar domínios existentes se especificado
            if existing_domains_file:
                existing_domains = self.load_existing_domains(existing_domains_file)
                original_count = len(all_domains)
                all_domains.update(existing_domains)
                self.logger.info(f"📝 Mesclados {len(existing_domains)} domínios existentes")
                self.logger.info(f"📊 Novos domínios encontrados: {original_count}")
            
            # Salvar resultados
            if all_domains:
                success = self.save_domains(all_domains, output_file)
                if success:
                    self.logger.info(f"📈 Processamento concluído:")
                    self.logger.info(f"   - Arquivos processados com sucesso: {successful_files}")
                    self.logger.info(f"   - Arquivos com erro: {failed_files}")
                    self.logger.info(f"   - Total de domínios únicos: {len(all_domains)}")
                    return True
            else:
                self.logger.warning("⚠️ Nenhum domínio foi extraído dos arquivos")
                return False
        
        except Exception as e:
            self.logger.error(f"❌ Erro durante o processamento: {e}")
            return False


def main():
    """Função principal"""
    # Configurações
    INPUT_FOLDER = r"C:\pdfs"
    OUTPUT_FILE = "dominios_extraidos.txt"
    EXISTING_DOMAINS_FILE = os.path.join(INPUT_FOLDER, "dominios.txt")
    
    # Criar extrator
    extractor = DomainExtractor(max_workers=8)
    
    # Executar extração
    success = extractor.extract_domains(
        input_folder=INPUT_FOLDER,
        output_file=OUTPUT_FILE,
        remove_duplicates=True,
        existing_domains_file=EXISTING_DOMAINS_FILE if os.path.exists(EXISTING_DOMAINS_FILE) else None
    )
    
    if success:
        print("\n🎉 Extração concluída com sucesso!")
    else:
        print("\n❌ Falha na extração de domínios")

if __name__ == "__main__":
    main()