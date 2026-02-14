import os
import re
import multiprocessing
from typing import Set, List
import logging

import fitz  # PyMuPDF
from openpyxl import load_workbook
from tqdm import tqdm

# Configurações de Validação de Domínio
MIN_DOMAIN_LENGTH = 4
MAX_DOMAIN_LENGTH = 253

# Regex otimizada para captura de domínios
DOMAIN_REGEX = re.compile(
    r'\b((?!-)[A-Za-z0-9-]{1,63}(?<!-)\.)+(?!-)[A-Za-z0-9-]{1,63}(?<!-)\b',
    re.IGNORECASE
)

# Configurar logging para exibir no terminal
logging.basicConfig(level=logging.INFO, format='%(message)s')

def validate_domain(domain: str) -> bool:
    """Valida um domínio seguindo regras ICANN."""
    if len(domain) < MIN_DOMAIN_LENGTH or len(domain) > MAX_DOMAIN_LENGTH:
        return False
    
    parts = domain.split('.')
    if len(parts) < 2:
        return False
    
    tld = parts[-1]
    if not tld.isalpha() or len(tld) < 2:
        return False
    
    for part in parts:
        if not (1 <= len(part) <= 63):
            return False
        if part.startswith('-') or part.endswith('-'):
            return False
        if not all(c.isalnum() or c == '-' for c in part):
            return False
    
    return True

def clean_domain(candidate: str) -> str | None:
    """Limpa e normaliza um domínio antes da validação."""
    candidate = candidate.split('//')[-1].split('/')[0].split('?')[0]
    candidate = candidate.split(':')[0].lower().strip('.:,;!*')
    
    if '.' not in candidate:
        return None
    
    parts = candidate.split('.')
    if not parts[-1].isalpha():
        return None
    
    return candidate if validate_domain(candidate) else None

def find_domains(text: str) -> Set[str]:
    """Procura e valida domínios no texto."""
    if not text:
        return set()
    
    found = set()
    for match in DOMAIN_REGEX.finditer(text):
        domain = clean_domain(match.group())
        if domain:
            found.add(domain)
    
    return found

def extract_text_from_pdf(pdf_path: str) -> str:
    """Extrai texto de PDFs."""
    try:
        with fitz.open(pdf_path) as doc:
            return " ".join(page.get_text("text") for page in doc)
    except Exception as e:
        logging.error(f"Erro no PDF {pdf_path}: {str(e)}")
        return ""

def extract_text_from_excel(excel_path: str) -> str:
    """Extrai texto de arquivos Excel."""
    try:
        texts = []
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        for sheet in wb:
            texts.extend(
                str(cell) for row in sheet.iter_rows(values_only=True)
                for cell in row if cell is not None
            )
        return "\n".join(texts)
    except Exception as e:
        logging.error(f"Erro no Excel {excel_path}: {str(e)}")
        return ""

def process_single_file(file_path: str) -> Set[str]:
    """Processa um único arquivo e extrai domínios."""
    try:
        extractors = {
            '.pdf': extract_text_from_pdf,
            '.xls': extract_text_from_excel,
            '.xlsx': extract_text_from_excel,
            '.ods': extract_text_from_excel
        }
        
        extractor = extractors.get(os.path.splitext(file_path)[1].lower())
        
        if not extractor:
            return set()
        
        text = extractor(file_path)
        domains = find_domains(text)
        
        if domains:
            logging.info(f"✅ Domínios encontrados em {os.path.basename(file_path)}: {len(domains)}")
        
        return domains
    
    except Exception as e:
        logging.error(f"Erro ao processar {os.path.basename(file_path)}: {str(e)}")
        return set()

def process_files(folder_path: str, batch_size: int = None, max_workers: int = None) -> Set[str]:
    """Processa arquivos em paralelo."""
    all_files = [
        os.path.join(root, f)
        for root, _, files in os.walk(folder_path)
        for f in files if os.path.splitext(f)[1].lower() in {'.pdf', '.xls', '.xlsx', '.ods'}
    ]
    
    max_workers = max_workers or (os.cpu_count() or 1)
    if batch_size is None:
        batch_size = max(1, len(all_files) // (max_workers * 2))
    
    all_domains = set()
    
    with multiprocessing.Pool(processes=max_workers) as pool:
        for batch_result in tqdm(
            pool.imap_unordered(process_single_file, all_files, chunksize=batch_size),
            total=len(all_files),
            desc="Processando arquivos"
        ):
            all_domains.update(batch_result)
    
    return all_domains

def read_existing_domains(file_path: str) -> Set[str]:
    """Lê domínios de um arquivo de texto."""
    if not os.path.exists(file_path):
        logging.warning(f"⚠️ Arquivo {file_path} não encontrado. Nenhum domínio existente será considerado.")
        return set()
    
    with open(file_path, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())

def main(input_folder: str = r"C:\pdfs", output_file: str = "dominios_extraidos.txt"):
    """Função principal para extração, combinação e salvamento de domínios."""
    logging.info("Iniciando processamento paralelo...")
    
    try:
        # Processar arquivos e extrair domínios
        extracted_domains = process_files(input_folder)
        
        # Ler domínios existentes do arquivo dominios.txt
        existing_domains_path = os.path.join(input_folder, "dominios.txt")
        existing_domains = read_existing_domains(existing_domains_path)
        
        # Verificar duplicatas entre extraídos e existentes
        duplicated_domains = [domain for domain in extracted_domains if domain in existing_domains]
        
        '''if duplicated_domains:
            logging.warning(f"\n⚠️ Aviso: {len(duplicated_domains)} domínios duplicados encontrados em relação ao dominios.txt:")
            for domain in duplicated_domains:
                logging.warning(f"  - {domain}")
        else:
            logging.info("\n✅ Nenhum domínio duplicado encontrado em relação ao dominios.txt.")
        '''
        
        # Combinar domínios extraídos e existentes, eliminando duplicatas
        combined_domains = sorted(extracted_domains.union(existing_domains))
        
        # Salvar a lista combinada no arquivo de saída
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(combined_domains))
        
        logging.info(f"\n✅ Concluído! Total de domínios únicos salvos: {len(combined_domains)}")
    
    except Exception as e:
        logging.error(f"Erro ao processar ou salvar resultados: {str(e)}")

if __name__ == '__main__':
    main()