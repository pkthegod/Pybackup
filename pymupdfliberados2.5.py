import os
import re
import fitz
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
import csv
import logging
from tqdm import tqdm

# Configurações
INPUT_FOLDER = r"C:\pdfs\liberados"
OUTPUT_FILE = "dominios_liberados.txt"
EXTENSOES = ('.pdf', '.xls', '.xlsx', '.ods', '.csv')
POPPLER_PATH = r"C:\poppler\Library\bin"  # Ajuste conforme sua instalação
TESSERACT_CMD = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Lista personalizada de domínios (adicione os seus aqui)
DOMINIOS_PERSONALIZADOS = {
    "blaze.com.br",
    "mpsp.mp.br",
    "policiacivil.pe.gov.br",
    "fazenda.gov.br",
    "anatel.gov.br",
}

# Configurar caminho do Tesseract
pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

logging.basicConfig(level=logging.INFO, format='%(message)s')

def extract_text_with_ocr(pdf_path: str) -> str:
    """Extrai texto de PDFs usando OCR"""
    try:
        images = convert_from_path(pdf_path, poppler_path=POPPLER_PATH, dpi=300)
        return "\n".join(pytesseract.image_to_string(img, lang='por') for img in images)
    except Exception as e:
        logging.error(f"Erro no OCR {pdf_path}: {str(e)}")
        return ""

def extract_text_from_pdf(pdf_path: str) -> str:
    """Combina extração direta com OCR"""
    try:
        text = ""
        with fitz.open(pdf_path) as doc:
            for page in doc:
                text += page.get_text("text", flags=fitz.TEXT_PRESERVE_WHITESPACE) + "\n"
        
        if len(text.strip()) > 50:
            return text
        
        logging.warning(f"Usando OCR para: {os.path.basename(pdf_path)}")
        return extract_text_with_ocr(pdf_path)
    except Exception as e:
        logging.error(f"Erro no PDF {pdf_path}: {str(e)}")
        return ""

def extract_text_from_excel(excel_path: str) -> str:
    """Extrai texto de planilhas Excel"""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        return "\n".join(" ".join(str(cell) for cell in row if cell) 
                       for sheet in wb for row in sheet.iter_rows(values_only=True))
    except Exception as e:
        logging.error(f"Erro no Excel {excel_path}: {str(e)}")
        return ""

def extract_text_from_csv(csv_path: str) -> str:
    """Extrai texto de arquivos CSV"""
    try:
        with open(csv_path, 'r', encoding='utf-8', errors='ignore') as f:
            return "\n".join(",".join(row) for row in csv.reader(f))
    except Exception as e:
        logging.error(f"Erro no CSV {csv_path}: {str(e)}")
        return ""

def find_domains(text: str) -> set:
    """Busca dominios com validação rigorosa"""
    pattern = r'(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,63}'
    matches = re.finditer(pattern, text)
    
    domains = set()
    for match in matches:
        domain = match.group().lower().strip('.:,;!*')
        if 4 <= len(domain) <= 253 and not domain.startswith(('.', '-')) and not domain.endswith(('.', '-')):
            domains.add(domain)
    
    return domains

def process_files():
    """Processa todos os arquivos e combina com a lista personalizada"""
    all_files = [os.path.join(root, f) 
                for root, _, files in os.walk(INPUT_FOLDER) 
                for f in files if f.lower().endswith(EXTENSOES)]
    
    if not all_files:
        logging.warning("Nenhum arquivo compatível encontrado!")
        return
    
    all_domains = set()
    
    for file_path in tqdm(all_files, desc="Processando arquivos"):
        try:
            if file_path.lower().endswith('.pdf'):
                text = extract_text_from_pdf(file_path)
            elif file_path.lower().endswith(('.xls', '.xlsx', '.ods')):
                text = extract_text_from_excel(file_path)
            elif file_path.lower().endswith('.csv'):
                text = extract_text_from_csv(file_path)
            
            if text:
                all_domains.update(find_domains(text))
                
        except Exception as e:
            logging.error(f"Erro em {os.path.basename(file_path)}: {str(e)}")
    
    # Adiciona lista personalizada e valida
    all_domains.update(DOMINIOS_PERSONALIZADOS)
    valid_domains = {d for d in all_domains if validate_domain(d)}
    
    # Salva resultados
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write("\n".join(sorted(valid_domains)))
    
    logging.info(f"\n✅ Processo concluído! Total de domínios: {len(valid_domains)}")
    logging.info(f"📌 Domínios personalizados adicionados: {len(DOMINIOS_PERSONALIZADOS)}")

def validate_domain(domain: str) -> bool:
    """Validação final do domínio"""
    parts = domain.split('.')
    return (len(parts) >= 2 and 
            all(1 <= len(part) <= 63 for part in parts) and 
            parts[-1].isalpha() and 
            not any(part.startswith('-') or part.endswith('-') for part in parts))

if __name__ == '__main__':
    process_files()